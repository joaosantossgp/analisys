from __future__ import annotations

import io
import logging
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi.testclient import TestClient

from apps.api.app.main import create_app
from src.company_catalog import CompanyCatalogEntry
from src.settings import build_settings


class StaticCompanyCatalog:
    def __init__(self, *entries: CompanyCatalogEntry):
        self.entries = tuple(entries)
        self.by_cd_cvm = {entry.cd_cvm: entry for entry in entries}

    def lookup_company(self, cd_cvm: int) -> CompanyCatalogEntry | None:
        return self.by_cd_cvm.get(int(cd_cvm))

    def search_companies(
        self,
        *,
        q: str,
        limit: int,
        exclude_codes=None,
    ) -> tuple[CompanyCatalogEntry, ...]:
        normalized = str(q or "").strip().lower()
        excluded = set(exclude_codes or set())
        if not normalized:
            return ()

        matches = []
        for entry in self.entries:
            if entry.cd_cvm in excluded:
                continue
            haystacks = [
                entry.company_name.lower(),
                (entry.nome_comercial or "").lower(),
                (entry.ticker_b3 or "").lower(),
                str(entry.cd_cvm),
            ]
            if any(normalized in value for value in haystacks):
                matches.append(entry)
        return tuple(matches[: int(limit)])


def test_health_returns_ok_payload(client: TestClient):
    response = client.get("/health")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["version"] == "v2-phase1"
    assert payload["database_dialect"] == client.app.state.read_service.engine.dialect.name
    assert payload["required_tables"] == ["financial_reports", "companies"]


def test_successful_healthcheck_does_not_emit_request_log(
    client: TestClient,
    caplog: pytest.LogCaptureFixture,
):
    caplog.set_level(logging.INFO, logger="cvm.api")
    caplog.clear()

    response = client.get("/health")

    assert response.status_code == 200
    request_logs = [
        record
        for record in caplog.records
        if record.name == "cvm.api" and "request_completed" in record.getMessage()
    ]
    assert request_logs == []


def test_degraded_healthcheck_emits_warning_request_log(
    client: TestClient,
    caplog: pytest.LogCaptureFixture,
    monkeypatch: pytest.MonkeyPatch,
):
    import apps.api.app.routes.health as health_route

    caplog.set_level(logging.INFO, logger="cvm.api")
    caplog.clear()
    monkeypatch.setattr(
        health_route,
        "collect_api_startup_report",
        lambda settings, warn_on_legacy_data=True: SimpleNamespace(ok=False, warnings=[], errors=[]),
    )

    response = client.get("/health")

    assert response.status_code == 503
    warning_logs = [
        record
        for record in caplog.records
        if record.name == "cvm.api"
        and record.levelno == logging.WARNING
        and "request_completed method=GET path=/health status=503" in record.getMessage()
    ]
    assert len(warning_logs) == 1


def test_application_requests_still_emit_info_request_log(
    client: TestClient,
    caplog: pytest.LogCaptureFixture,
):
    caplog.set_level(logging.INFO, logger="cvm.api")
    caplog.clear()

    response = client.get("/companies")

    assert response.status_code == 200
    info_logs = [
        record
        for record in caplog.records
        if record.name == "cvm.api"
        and record.levelno == logging.INFO
        and "request_completed method=GET path=/companies status=200" in record.getMessage()
    ]
    assert len(info_logs) == 1


def test_companies_empty_search_returns_paginated_directory(client: TestClient):
    response = client.get("/companies")

    assert response.status_code == 200
    payload = response.json()
    assert payload["pagination"]["total_items"] == 4
    assert payload["pagination"]["page"] == 1
    assert payload["pagination"]["page_size"] == 20
    assert payload["items"][0]["company_name"] == "PETROBRAS"
    assert payload["items"][0]["anos_disponiveis"] == [2023, 2024]
    assert payload["items"][0]["sector_name"] == "Energia"
    assert payload["items"][0]["sector_slug"] == "energia"
    assert payload["items"][0]["has_financial_data"] is True
    returned_names = [item["company_name"] for item in payload["items"]]
    assert "SEM DADOS" in returned_names
    sem_dados = next(i for i in payload["items"] if i["company_name"] == "SEM DADOS")
    assert sem_dados["has_financial_data"] is False
    assert sem_dados["anos_disponiveis"] == []


@pytest.mark.parametrize(
    ("path", "params", "expected_cache_control"),
    [
        ("/companies", None, "public, max-age=300, stale-while-revalidate=3600"),
        ("/companies/9512", None, "public, max-age=3600"),
        ("/companies/9512/years", None, "public, max-age=86400, stale-while-revalidate=604800"),
        (
            "/companies/9512/statements",
            {"stmt": "DRE", "years": "2023,2024"},
            "public, max-age=600",
        ),
        (
            "/companies/9512/kpis",
            {"years": "2023,2024"},
            "public, max-age=600",
        ),
        ("/sectors", None, "public, max-age=3600, stale-while-revalidate=86400"),
        ("/sectors/energia", None, "public, max-age=3600, stale-while-revalidate=86400"),
        ("/companies/filters", None, "public, max-age=3600, stale-while-revalidate=86400"),
        ("/companies/suggestions", None, "public, max-age=60, stale-while-revalidate=300"),
        ("/companies/suggestions", {"q": "petro"}, "public, max-age=60, stale-while-revalidate=300"),
    ],
)
def test_cacheable_endpoints_expose_cache_headers(
    client: TestClient,
    path: str,
    params: dict[str, str] | None,
    expected_cache_control: str,
):
    response = client.get(path, params=params)

    assert response.status_code == 200
    assert response.headers["cache-control"] == expected_cache_control
    vary_values = {value.strip() for value in response.headers["vary"].split(",")}
    assert "Origin" in vary_values


def test_companies_search_filters_results(client: TestClient):
    response = client.get("/companies", params={"search": "vale", "page_size": 20})

    assert response.status_code == 200
    payload = response.json()
    assert payload["pagination"]["total_items"] == 1
    assert len(payload["items"]) == 1
    assert payload["items"][0]["company_name"] == "VALE"


def test_companies_pagination_respects_page_and_page_size(client: TestClient):
    response = client.get("/companies", params={"page": 2, "page_size": 1})

    assert response.status_code == 200
    payload = response.json()
    assert payload["pagination"] == {
        "page": 2,
        "page_size": 1,
        "total_items": 4,
        "total_pages": 4,
        "has_next": True,
        "has_previous": True,
    }
    assert len(payload["items"]) == 1
    assert payload["items"][0]["company_name"] == "SABESP"


def test_companies_sector_filter_uses_canonical_slug(client: TestClient):
    response = client.get("/companies", params={"sector": "saneamento"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["applied_filters"]["sector"] == "saneamento"
    assert payload["pagination"]["total_items"] == 1
    assert payload["items"][0]["company_name"] == "SABESP"
    assert payload["items"][0]["sector_name"] == "Saneamento"


def test_companies_unknown_sector_returns_empty_page_with_stable_payload(client: TestClient):
    response = client.get("/companies", params={"sector": "setor-inexistente"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["items"] == []
    assert payload["pagination"] == {
        "page": 1,
        "page_size": 20,
        "total_items": 0,
        "total_pages": 1,
        "has_next": False,
        "has_previous": False,
    }
    assert payload["applied_filters"] == {"search": "", "sector": "setor-inexistente"}


def test_companies_filters_returns_canonical_sector_options(client: TestClient):
    response = client.get("/companies/filters")

    assert response.status_code == 200
    payload = response.json()
    assert payload["sectors"] == [
        {"sector_name": "Energia", "sector_slug": "energia", "company_count": 1},
        {"sector_name": "Financeiro", "sector_slug": "financeiro", "company_count": 1},
        {"sector_name": "Materiais Basicos", "sector_slug": "materiais-basicos", "company_count": 1},
        {"sector_name": "Saneamento", "sector_slug": "saneamento", "company_count": 1},
    ]


def test_company_suggestions_empty_query_returns_items_alphabetically(client: TestClient):
    response = client.get("/companies/suggestions")

    assert response.status_code == 200
    payload = response.json()
    assert "items" in payload
    names = [item["company_name"] for item in payload["items"]]
    assert names == sorted(names)


def test_company_suggestions_payload_has_minimal_fields_only(client: TestClient):
    response = client.get("/companies/suggestions", params={"q": "petro"})

    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) == 1
    item = items[0]
    assert set(item.keys()) == {"cd_cvm", "company_name", "ticker_b3", "sector_slug"}
    assert item["cd_cvm"] == 9512
    assert item["company_name"] == "PETROBRAS"
    assert item["ticker_b3"] == "PETR4"
    assert item["sector_slug"] == "energia"


def test_company_suggestions_filters_by_ticker(client: TestClient):
    response = client.get("/companies/suggestions", params={"q": "vale3"})

    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) == 1
    assert items[0]["company_name"] == "VALE"


def test_company_suggestions_exact_ticker_ranks_first(client: TestClient):
    response = client.get("/companies/suggestions", params={"q": "petr4"})

    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) >= 1
    assert items[0]["cd_cvm"] == 9512


def test_company_suggestions_respects_limit(client: TestClient):
    response = client.get("/companies/suggestions", params={"limit": 2})

    assert response.status_code == 200
    assert len(response.json()["items"]) <= 2


def test_company_suggestions_returns_empty_for_no_match(client: TestClient):
    response = client.get("/companies/suggestions", params={"q": "xxxxnomatch"})

    assert response.status_code == 200
    assert response.json()["items"] == []


def test_company_suggestions_fall_back_to_catalog_when_local_matches_are_missing(
    client: TestClient,
):
    client.app.state.read_service._company_catalog = StaticCompanyCatalog(
        CompanyCatalogEntry(
            cd_cvm=19348,
            company_name="ITAU UNIBANCO HOLDING S.A.",
            nome_comercial="Itau Unibanco",
            cnpj="60.872.504/0001-23",
            setor_cvm="Financeiro",
            ticker_b3="ITUB4",
            is_active=True,
        )
    )

    response = client.get("/companies/suggestions", params={"q": "itub4"})

    assert response.status_code == 200
    assert response.json()["items"] == [
        {
            "cd_cvm": 19348,
            "company_name": "ITAU UNIBANCO HOLDING S.A.",
            "ticker_b3": "ITUB4",
            "sector_slug": "financeiro",
        }
    ]


def test_company_suggestions_rejects_limit_above_max(client: TestClient):
    response = client.get("/companies/suggestions", params={"limit": 21})

    assert response.status_code == 422


def test_company_suggestions_rejects_limit_below_min(client: TestClient):
    response = client.get("/companies/suggestions", params={"limit": 0})

    assert response.status_code == 422


def test_sectors_returns_directory_with_latest_year_and_snapshot(client: TestClient):
    response = client.get("/sectors")

    assert response.status_code == 200
    payload = response.json()
    assert [item["sector_name"] for item in payload["items"]] == [
        "Energia",
        "Materiais Basicos",
        "Saneamento",
    ]
    energia = payload["items"][0]
    assert energia["sector_slug"] == "energia"
    assert energia["company_count"] == 1
    assert energia["latest_year"] == 2024
    assert energia["snapshot"]["roe"] == pytest.approx(180.0 / 380.0)
    assert energia["snapshot"]["mg_ebit"] == pytest.approx(240.0 / 1100.0)
    assert energia["snapshot"]["mg_liq"] == pytest.approx(180.0 / 1100.0)


def test_sectors_directory_keeps_null_snapshot_metrics_when_accounts_are_partial(client: TestClient):
    response = client.get("/sectors")

    assert response.status_code == 200
    payload = response.json()
    materiais = next(item for item in payload["items"] if item["sector_slug"] == "materiais-basicos")
    assert materiais["latest_year"] == 2024
    assert materiais["snapshot"] == {"roe": None, "mg_ebit": None, "mg_liq": None}


def test_sector_detail_returns_default_latest_year_and_yearly_overview(client: TestClient):
    response = client.get("/sectors/energia")

    assert response.status_code == 200
    payload = response.json()
    assert payload["sector_name"] == "Energia"
    assert payload["sector_slug"] == "energia"
    assert payload["company_count"] == 1
    assert payload["available_years"] == [2023, 2024]
    assert payload["selected_year"] == 2024
    assert payload["yearly_overview"] == [
        {
            "year": 2023,
            "roe": pytest.approx(150.0 / 300.0),
            "mg_ebit": pytest.approx(200.0 / 1000.0),
            "mg_liq": pytest.approx(150.0 / 1000.0),
        },
        {
            "year": 2024,
            "roe": pytest.approx(180.0 / 380.0),
            "mg_ebit": pytest.approx(240.0 / 1100.0),
            "mg_liq": pytest.approx(180.0 / 1100.0),
        },
    ]
    assert payload["companies"] == [
        {
            "cd_cvm": 9512,
            "company_name": "PETROBRAS",
            "ticker_b3": "PETR4",
            "roe": pytest.approx(180.0 / 380.0),
            "mg_ebit": pytest.approx(240.0 / 1100.0),
            "mg_liq": pytest.approx(180.0 / 1100.0),
        }
    ]


def test_sector_detail_respects_explicit_year(client: TestClient):
    response = client.get("/sectors/energia", params={"year": "2023"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["selected_year"] == 2023
    assert payload["companies"][0]["roe"] == pytest.approx(150.0 / 300.0)


def test_sector_detail_keeps_company_row_with_null_metrics_when_accounts_are_partial(client: TestClient):
    response = client.get("/sectors/saneamento")

    assert response.status_code == 200
    payload = response.json()
    assert payload["selected_year"] == 2024
    assert payload["companies"] == [
        {
            "cd_cvm": 11223,
            "company_name": "SABESP",
            "ticker_b3": "SBSP3",
            "roe": None,
            "mg_ebit": None,
            "mg_liq": None,
        }
    ]


def test_request_refresh_returns_202_and_enqueues_internal_job(client: TestClient):
    from sqlalchemy import text as sa_text

    response = client.post("/companies/9512/request-refresh")

    assert response.status_code == 202
    payload = response.json()
    assert payload["cd_cvm"] == 9512
    assert payload["status"] == "queued"
    assert isinstance(payload["job_id"], str) and payload["job_id"]
    assert isinstance(payload["accepted_at"], str)
    assert "enfileirada" in payload["message"].lower()

    engine = client.app.state.read_service.engine
    with engine.connect() as conn:
        job_row = conn.execute(
            sa_text(
                """
                SELECT state, stage, source_scope, cd_cvm
                FROM refresh_jobs
                WHERE id = :job_id
                """
            ),
            {"job_id": payload["job_id"]},
        ).mappings().one()
        projection_row = conn.execute(
            sa_text(
                """
                SELECT last_status, source_scope, job_id, queue_position
                FROM company_refresh_status
                WHERE cd_cvm = 9512
                """
            )
        ).mappings().one()

    assert job_row["state"] == "queued"
    assert job_row["stage"] == "queued"
    assert job_row["source_scope"] == "on_demand"
    assert int(job_row["cd_cvm"]) == 9512
    assert projection_row["last_status"] == "queued"
    assert projection_row["source_scope"] == "on_demand"
    assert projection_row["job_id"] == payload["job_id"]
    assert projection_row["queue_position"] == 0


def test_request_refresh_bootstraps_unknown_local_company_from_catalog(
    client: TestClient,
):
    from sqlalchemy import text as sa_text

    client.app.state.read_service._company_catalog = StaticCompanyCatalog(
        CompanyCatalogEntry(
            cd_cvm=19348,
            company_name="ITAU UNIBANCO HOLDING S.A.",
            nome_comercial="Itau Unibanco",
            cnpj="60.872.504/0001-23",
            setor_cvm="Financeiro",
            ticker_b3="ITUB4",
            is_active=True,
        )
    )

    response = client.post("/companies/19348/request-refresh")

    assert response.status_code == 202
    payload = response.json()
    assert payload["cd_cvm"] == 19348
    assert payload["status"] == "queued"
    assert isinstance(payload["job_id"], str) and payload["job_id"]

    engine = client.app.state.read_service.engine
    with engine.connect() as conn:
        company_row = conn.execute(
            sa_text(
                """
                SELECT company_name, nome_comercial, setor_cvm, ticker_b3
                FROM companies
                WHERE cd_cvm = 19348
                """
            )
        ).mappings().one()
        refresh_row = conn.execute(
            sa_text(
                """
                SELECT source_scope, last_status, job_id
                FROM company_refresh_status
                WHERE cd_cvm = 19348
                """
            )
        ).mappings().one()

    assert company_row["company_name"] == "ITAU UNIBANCO HOLDING S.A."
    assert company_row["nome_comercial"] == "Itau Unibanco"
    assert company_row["setor_cvm"] == "Financeiro"
    assert company_row["ticker_b3"] == "ITUB4"
    assert refresh_row["source_scope"] == "on_demand_bootstrap"
    assert refresh_row["last_status"] == "queued"
    assert refresh_row["job_id"] == payload["job_id"]


def test_request_refresh_returns_404_for_unknown_company(client: TestClient):
    response = client.post("/companies/9999999/request-refresh")

    assert response.status_code == 404


def test_request_refresh_returns_already_current_without_creating_job(
    client: TestClient,
):
    from sqlalchemy import text as sa_text

    engine = client.app.state.read_service.engine
    with engine.begin() as conn:
        annual_rows = []
        for report_year in range(2010, 2026):
            for statement_type in ("BPA", "BPP", "DRE", "DFC"):
                annual_rows.append(
                    {
                        "COMPANY_NAME": "VALE",
                        "CD_CVM": 4170,
                        "STATEMENT_TYPE": statement_type,
                        "REPORT_YEAR": report_year,
                        "PERIOD_LABEL": str(report_year),
                        "LINE_ID_BASE": f"{statement_type.lower()}-{report_year}",
                        "CD_CONTA": statement_type,
                        "DS_CONTA": statement_type,
                        "STANDARD_NAME": statement_type,
                        "QA_CONFLICT": 0,
                        "VL_CONTA": 1.0,
                    }
                )
        conn.execute(
            sa_text(
                """
                INSERT INTO financial_reports (
                    COMPANY_NAME, CD_CVM, STATEMENT_TYPE, REPORT_YEAR, PERIOD_LABEL,
                    LINE_ID_BASE, CD_CONTA, DS_CONTA, STANDARD_NAME, QA_CONFLICT, VL_CONTA
                ) VALUES (
                    :COMPANY_NAME, :CD_CVM, :STATEMENT_TYPE, :REPORT_YEAR, :PERIOD_LABEL,
                    :LINE_ID_BASE, :CD_CONTA, :DS_CONTA, :STANDARD_NAME, :QA_CONFLICT, :VL_CONTA
                )
                """
            ),
            annual_rows,
        )

    response = client.post("/companies/4170/request-refresh")

    assert response.status_code == 202
    payload = response.json()
    assert payload["status"] == "already_current"
    assert payload["job_id"] is None
    assert "ja atualizada" in payload["message"].lower()

    with engine.connect() as conn:
        queued_jobs = conn.execute(
            sa_text("SELECT COUNT(*) FROM refresh_jobs WHERE cd_cvm = 4170")
        ).scalar_one()
        projection_row = conn.execute(
            sa_text(
                """
                SELECT last_status, job_id, progress_message
                FROM company_refresh_status
                WHERE cd_cvm = 4170
                """
            )
        ).mappings().one()

    assert int(queued_jobs) == 0
    assert projection_row["last_status"] == "success"
    assert projection_row["job_id"] is None
    assert "ja atualizada" in str(projection_row["progress_message"]).lower()


@pytest.mark.parametrize("state", ["queued", "running"])
def test_request_refresh_returns_429_when_active_job_exists(
    client: TestClient,
    state: str,
):
    from sqlalchemy import text as sa_text

    engine = client.app.state.read_service.engine
    with engine.begin() as conn:
        conn.execute(
            sa_text(
                """
                INSERT INTO refresh_jobs (
                    id, cd_cvm, company_name, source_scope,
                    start_year, end_year, state, stage, requested_at
                ) VALUES (
                    :id, 4170, 'VALE', 'on_demand',
                    2010, 2025, :state, :stage, '2026-04-21T12:00:00+00:00'
                )
                """
            ),
            {
                "id": f"active-{state}",
                "state": state,
                "stage": "planning" if state == "running" else "queued",
            },
        )

    response = client.post("/companies/4170/request-refresh")

    assert response.status_code == 429
    assert response.json()["detail"]["code"] == "refresh_already_active"


def test_request_top_ranked_historical_refresh_queues_only_missing_and_excludes_no_data(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
):
    from sqlalchemy import text as sa_text
    import src.github_dispatch as github_dispatch

    calls: list[tuple[int, int | None, int | None]] = []

    def fake_dispatch(
        cd_cvm: int,
        *,
        start_year: int | None = None,
        end_year: int | None = None,
    ) -> tuple[bool, str | None]:
        calls.append((cd_cvm, start_year, end_year))
        return True, None

    monkeypatch.setattr(github_dispatch, "dispatch_on_demand_ingest", fake_dispatch)

    engine = client.app.state.read_service.engine
    with engine.begin() as conn:
        conn.execute(
            sa_text(
                """
                INSERT INTO company_refresh_status (
                    cd_cvm,
                    company_name,
                    source_scope,
                    last_status,
                    last_attempt_at,
                    updated_at
                ) VALUES (
                    11223,
                    'SABESP',
                    'local',
                    'no_data',
                    '2026-04-08T09:00:00',
                    '2026-04-08T09:00:00'
                )
                """
            )
        )

    response = client.post(
        "/companies/request-refresh/top-ranked",
        params={"limit": 3, "start_year": 2023, "end_year": 2024},
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["total_ranked"] == 3
    assert payload["queued_count"] == 1
    assert payload["already_queued_count"] == 0
    assert payload["no_data_excluded_count"] == 1
    assert payload["already_complete_count"] == 1
    assert payload["dispatch_failed_count"] == 0
    assert calls == [(4170, 2023, 2024)]

    items_by_company = {item["cd_cvm"]: item for item in payload["items"]}
    assert items_by_company[9512]["status"] == "already_complete"
    assert items_by_company[4170]["status"] == "queued"
    assert items_by_company[4170]["years_missing"] == [2023]
    assert items_by_company[11223]["status"] == "no_data_excluded"

    with engine.connect() as conn:
        row = conn.execute(
            sa_text(
                """
                SELECT source_scope, last_status, last_start_year, last_end_year
                FROM company_refresh_status
                WHERE cd_cvm = 4170
                """
            )
        ).mappings().one()

    assert row["source_scope"] == "ranked_backfill"
    assert row["last_status"] == "queued"
    assert row["last_start_year"] == 2023
    assert row["last_end_year"] == 2024


def test_sector_detail_returns_404_for_unknown_slug(client: TestClient):
    response = client.get("/sectors/setor-inexistente")

    assert response.status_code == 404
    assert response.json()["error"]["code"] == "not_found"


def test_sector_detail_returns_422_for_year_outside_available_range(client: TestClient):
    response = client.get("/sectors/energia", params={"year": "1990"})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_detail_returns_metadata(client: TestClient):
    response = client.get("/companies/9512")

    assert response.status_code == 200
    payload = response.json()
    assert payload["company_name"] == "PETROBRAS"
    assert payload["ticker_b3"] == "PETR4"
    assert payload["sector_name"] == "Energia"
    assert payload["sector_slug"] == "energia"


def test_company_detail_uses_catalog_fallback_for_unknown_local_company(
    client: TestClient,
):
    client.app.state.read_service._company_catalog = StaticCompanyCatalog(
        CompanyCatalogEntry(
            cd_cvm=19348,
            company_name="ITAU UNIBANCO HOLDING S.A.",
            nome_comercial="Itau Unibanco",
            cnpj="60.872.504/0001-23",
            setor_cvm="Financeiro",
            ticker_b3="ITUB4",
            is_active=True,
        )
    )

    response = client.get("/companies/19348")

    assert response.status_code == 200
    payload = response.json()
    assert payload["company_name"] == "ITAU UNIBANCO HOLDING S.A."
    assert payload["nome_comercial"] == "Itau Unibanco"
    assert payload["sector_name"] == "Financeiro"
    assert payload["sector_slug"] == "financeiro"
    assert payload["ticker_b3"] == "ITUB4"


def test_company_detail_uses_sector_fallback_when_analytical_sector_is_missing(client: TestClient):
    response = client.get("/companies/11223")

    assert response.status_code == 200
    payload = response.json()
    assert payload["company_name"] == "SABESP"
    assert payload["sector_name"] == "Saneamento"
    assert payload["sector_slug"] == "saneamento"


def test_company_detail_returns_404_for_unknown_company(client: TestClient):
    response = client.get("/companies/999999")

    assert response.status_code == 404
    assert response.json()["error"]["code"] == "not_found"


def test_company_years_returns_sorted_values(client: TestClient):
    response = client.get("/companies/9512/years")

    assert response.status_code == 200
    assert response.json() == [2023, 2024]


def test_company_years_returns_empty_list_when_company_has_no_reports(client: TestClient):
    response = client.get("/companies/77889/years")

    assert response.status_code == 200
    assert response.json() == []


def test_company_years_return_empty_list_for_catalog_company_without_local_reports(
    client: TestClient,
):
    client.app.state.read_service._company_catalog = StaticCompanyCatalog(
        CompanyCatalogEntry(
            cd_cvm=19348,
            company_name="ITAU UNIBANCO HOLDING S.A.",
            nome_comercial="Itau Unibanco",
            cnpj="60.872.504/0001-23",
            setor_cvm="Financeiro",
            ticker_b3="ITUB4",
            is_active=True,
        )
    )

    response = client.get("/companies/19348/years")

    assert response.status_code == 200
    assert response.json() == []


def test_company_excel_export_returns_binary_workbook(client: TestClient):
    response = client.get("/companies/9512/export/excel")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="PETR4_' in response.headers["content-disposition"]

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames[:4] == ["CAPA", "GERAL", "KPIs", "DRE"]


def test_company_excel_export_uses_all_available_years(client: TestClient):
    response = client.get("/companies/9512/export/excel")

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    dre = workbook["DRE"]

    assert dre["E1"].value == "2023"
    assert dre["F1"].value == "2024"


def test_company_excel_export_returns_404_for_unknown_company(client: TestClient):
    response = client.get("/companies/999999/export/excel")

    assert response.status_code == 404
    assert response.json()["error"]["code"] == "not_found"


def test_company_excel_export_returns_422_when_company_has_no_exportable_years(client: TestClient):
    response = client.get("/companies/77889/export/excel")

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_excel_batch_export_returns_zip_with_one_workbook_per_company(client: TestClient):
    response = client.get("/companies/export/excel-batch", params={"ids": "9512,4170"})

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/zip")
    assert 'filename="comparar_excel_lote.zip"' in response.headers["content-disposition"]

    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        names = sorted(archive.namelist())
        assert len(names) == 2
        assert names[0].startswith("PETR4_")
        assert names[1].startswith("VALE3_")
        assert names[0].endswith(".xlsx")
        assert names[1].endswith(".xlsx")

        workbook = openpyxl.load_workbook(io.BytesIO(archive.read(names[0])))
        assert "CAPA" in workbook.sheetnames


def test_company_excel_batch_export_rejects_duplicate_ids(client: TestClient):
    response = client.get("/companies/export/excel-batch", params={"ids": "9512,9512"})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_excel_batch_export_rejects_single_company(client: TestClient):
    response = client.get("/companies/export/excel-batch", params={"ids": "9512"})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_excel_batch_export_returns_404_for_unknown_company(client: TestClient):
    response = client.get("/companies/export/excel-batch", params={"ids": "9512,999999"})

    assert response.status_code == 404
    assert response.json()["error"]["code"] == "not_found"


def test_company_statement_returns_matrix(client: TestClient):
    response = client.get(
        "/companies/9512/statements",
        params={"stmt": "DRE", "years": "2023,2024"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["statement_type"] == "DRE"
    assert payload["years"] == [2023, 2024]
    assert "2023" in payload["table"]["columns"]
    assert "2024" in payload["table"]["columns"]


def test_company_statement_sorts_requested_years_for_stable_contract(client: TestClient):
    response = client.get(
        "/companies/9512/statements",
        params={"stmt": "DRE", "years": "2024,2023"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["years"] == [2023, 2024]
    assert payload["table"]["columns"][-2:] == ["2023", "2024"]


def test_company_statement_returns_empty_table_for_year_without_data(client: TestClient):
    response = client.get(
        "/companies/4170/statements",
        params={"stmt": "DRE", "years": "2023"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["years"] == [2023]
    assert payload["table"] == {"columns": [], "rows": []}


def test_company_statement_rejects_invalid_years(client: TestClient):
    response = client.get(
        "/companies/9512/statements",
        params={"stmt": "DRE", "years": "2024,foo"},
    )

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_statement_rejects_duplicate_years(client: TestClient):
    response = client.get(
        "/companies/9512/statements",
        params={"stmt": "DRE", "years": "2024,2024"},
    )

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_statement_rejects_invalid_statement(client: TestClient):
    response = client.get(
        "/companies/9512/statements",
        params={"stmt": "XYZ", "years": "2024"},
    )

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_kpis_returns_annual_and_quarterly_tables(client: TestClient):
    response = client.get("/companies/9512/kpis", params={"years": "2023,2024"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["cd_cvm"] == 9512
    assert payload["years"] == [2023, 2024]
    assert payload["annual"]["rows"]
    assert payload["quarterly"]["rows"]


def test_company_kpis_sort_requested_years_for_stable_contract(client: TestClient):
    response = client.get("/companies/9512/kpis", params={"years": "2024,2023"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["years"] == [2023, 2024]
    assert "2023" in payload["annual"]["columns"]
    assert "2024" in payload["annual"]["columns"]


def test_company_kpis_return_empty_tables_when_requested_year_has_no_data(client: TestClient):
    response = client.get("/companies/4170/kpis", params={"years": "2023"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["years"] == [2023]
    assert payload["annual"] == {"columns": [], "rows": []}
    assert payload["quarterly"] == {"columns": [], "rows": []}


def test_refresh_status_returns_operational_rows(client: TestClient):
    response = client.get("/refresh-status")

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["last_status"] == "success"
    assert payload[0]["job_id"] is None
    assert payload[0]["stage"] is None
    assert payload[0]["queue_position"] is None
    assert payload[0]["estimated_progress_pct"] is None
    assert payload[0]["estimated_eta_seconds"] is None


def test_refresh_status_exposes_terminal_no_data_state(client: TestClient):
    from sqlalchemy import text as sa_text

    with client.app.state.read_service.engine.begin() as conn:
        conn.execute(
            sa_text(
                """
                INSERT INTO company_refresh_status (
                    cd_cvm, company_name, source_scope, last_attempt_at, last_success_at,
                    last_status, last_error, last_start_year, last_end_year,
                    last_rows_inserted, updated_at, progress_message, finished_at
                ) VALUES (
                    4170, 'VALE', 'on_demand', '2026-04-21T12:00:00+00:00', NULL,
                    'no_data', NULL, 2010, 2025,
                    NULL, '2026-04-21T12:01:00+00:00', 'Nenhuma demonstracao encontrada para 2010-2025.',
                    '2026-04-21T12:01:00+00:00'
                )
                ON CONFLICT (cd_cvm) DO UPDATE SET
                    last_status = excluded.last_status,
                    progress_message = excluded.progress_message,
                    finished_at = excluded.finished_at,
                    updated_at = excluded.updated_at
                """
            )
        )

    response = client.get("/refresh-status", params={"cd_cvm": 4170})

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["last_status"] == "no_data"
    assert payload[0]["progress_message"] == "Nenhuma demonstracao encontrada para 2010-2025."
    assert payload[0]["finished_at"] == "2026-04-21T12:01:00+00:00"
    assert payload[0]["estimated_progress_pct"] is None


def test_refresh_status_returns_real_progress_fields_for_active_refresh(client: TestClient):
    from sqlalchemy import text as sa_text

    now = datetime.now(timezone.utc).replace(microsecond=0)
    now_iso = now.isoformat()
    started_at = (now - timedelta(minutes=4)).isoformat()

    with client.app.state.read_service.engine.begin() as conn:
        conn.execute(
            sa_text(
                """
                INSERT INTO company_refresh_status (
                    cd_cvm, company_name, source_scope, last_attempt_at, last_success_at,
                    last_status, last_error, last_start_year, last_end_year,
                    last_rows_inserted, updated_at, job_id, stage, queue_position,
                    progress_current, progress_total, progress_message,
                    started_at, heartbeat_at, finished_at
                ) VALUES (
                    :cd_cvm, :company_name, :source_scope, :last_attempt_at, :last_success_at,
                    :last_status, :last_error, :last_start_year, :last_end_year,
                    :last_rows_inserted, :updated_at, :job_id, :stage, :queue_position,
                    :progress_current, :progress_total, :progress_message,
                    :started_at, :heartbeat_at, :finished_at
                )
                ON CONFLICT (cd_cvm) DO UPDATE SET
                    company_name = excluded.company_name,
                    source_scope = excluded.source_scope,
                    last_attempt_at = excluded.last_attempt_at,
                    last_success_at = excluded.last_success_at,
                    last_status = excluded.last_status,
                    last_error = excluded.last_error,
                    last_start_year = excluded.last_start_year,
                    last_end_year = excluded.last_end_year,
                    last_rows_inserted = excluded.last_rows_inserted,
                    updated_at = excluded.updated_at,
                    job_id = excluded.job_id,
                    stage = excluded.stage,
                    queue_position = excluded.queue_position,
                    progress_current = excluded.progress_current,
                    progress_total = excluded.progress_total,
                    progress_message = excluded.progress_message,
                    started_at = excluded.started_at,
                    heartbeat_at = excluded.heartbeat_at,
                    finished_at = excluded.finished_at
                """
            ),
            [
                {
                    "cd_cvm": 4170,
                    "company_name": "VALE",
                    "source_scope": "on_demand",
                    "job_id": "job-active-vale",
                    "stage": "download_extract",
                    "queue_position": 1,
                    "last_attempt_at": started_at,
                    "last_success_at": None,
                    "last_status": "queued",
                    "last_error": None,
                    "last_start_year": 2010,
                    "last_end_year": 2024,
                    "last_rows_inserted": None,
                    "progress_current": 9,
                    "progress_total": 20,
                    "progress_message": "Download concluido para DFP/2018.",
                    "started_at": started_at,
                    "heartbeat_at": now_iso,
                    "finished_at": None,
                    "updated_at": now_iso,
                },
                {
                    "cd_cvm": 19348,
                    "company_name": "LOCALIZA",
                    "source_scope": "local",
                    "job_id": None,
                    "stage": None,
                    "queue_position": None,
                    "last_attempt_at": (now - timedelta(hours=3)).isoformat(),
                    "last_success_at": (now - timedelta(hours=2, minutes=44)).isoformat(),
                    "last_status": "success",
                    "last_error": None,
                    "last_start_year": 2010,
                    "last_end_year": 2024,
                    "last_rows_inserted": 120,
                    "progress_current": None,
                    "progress_total": None,
                    "progress_message": None,
                    "started_at": None,
                    "heartbeat_at": None,
                    "finished_at": None,
                    "updated_at": (now - timedelta(hours=2, minutes=44)).isoformat(),
                },
                {
                    "cd_cvm": 20532,
                    "company_name": "WEG",
                    "source_scope": "local",
                    "job_id": None,
                    "stage": None,
                    "queue_position": None,
                    "last_attempt_at": (now - timedelta(hours=2, minutes=30)).isoformat(),
                    "last_success_at": (now - timedelta(hours=2, minutes=12)).isoformat(),
                    "last_status": "success",
                    "last_error": None,
                    "last_start_year": 2010,
                    "last_end_year": 2024,
                    "last_rows_inserted": 116,
                    "progress_current": None,
                    "progress_total": None,
                    "progress_message": None,
                    "started_at": None,
                    "heartbeat_at": None,
                    "finished_at": None,
                    "updated_at": (now - timedelta(hours=2, minutes=12)).isoformat(),
                },
                {
                    "cd_cvm": 90678,
                    "company_name": "ITAU",
                    "source_scope": "local",
                    "job_id": None,
                    "stage": None,
                    "queue_position": None,
                    "last_attempt_at": (now - timedelta(hours=2)).isoformat(),
                    "last_success_at": (now - timedelta(hours=1, minutes=47)).isoformat(),
                    "last_status": "success",
                    "last_error": None,
                    "last_start_year": 2010,
                    "last_end_year": 2024,
                    "last_rows_inserted": 112,
                    "progress_current": None,
                    "progress_total": None,
                    "progress_message": None,
                    "started_at": None,
                    "heartbeat_at": None,
                    "finished_at": None,
                    "updated_at": (now - timedelta(hours=1, minutes=47)).isoformat(),
                },
            ],
        )

    response = client.get("/refresh-status", params={"cd_cvm": 4170})

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["last_status"] == "queued"
    assert payload[0]["job_id"] == "job-active-vale"
    assert payload[0]["stage"] == "download_extract"
    assert payload[0]["queue_position"] == 1
    assert payload[0]["progress_current"] == 9
    assert payload[0]["progress_total"] == 20
    assert payload[0]["progress_message"] == "Download concluido para DFP/2018."
    assert payload[0]["estimated_progress_pct"] == pytest.approx(25.3, abs=0.2)
    assert payload[0]["estimated_eta_seconds"] is not None
    assert payload[0]["estimated_total_seconds"] is not None
    assert payload[0]["elapsed_seconds"] >= 240
    assert payload[0]["estimated_completion_at"] is not None
    assert payload[0]["estimate_confidence"] == "high"


def test_base_health_returns_snapshot(client: TestClient):
    response = client.get("/base-health", params={"start_year": 2023, "end_year": 2024})

    assert response.status_code == 200
    payload = response.json()
    assert payload["start_year"] == 2023
    assert payload["end_year"] == 2024
    assert payload["total_cells"] == 8
    assert payload["completed_cells"] == 4
    assert payload["missing_cells"] == 4
    assert payload["pct"] == pytest.approx(50.0)
    assert payload["health_status"] in {"ok", "atencao", "critico"}
    assert payload["per_year"] == [
        {
            "year": 2023,
            "total_companies": 4,
            "completed": 1,
            "missing": 3,
            "pct": 25.0,
            "eta_hours": None,
        },
        {
            "year": 2024,
            "total_companies": 4,
            "completed": 3,
            "missing": 1,
            "pct": 75.0,
            "eta_hours": None,
        },
    ]
    assert payload["raw"]["ranked_backlog"]["total_ranked"] == 3
    assert payload["raw"]["ranked_backlog"]["fully_covered_companies"] == 1
    assert payload["raw"]["ranked_backlog"]["queue_eligible_companies"] == 2


def test_base_health_rejects_invalid_year_window(client: TestClient):
    response = client.get("/base-health", params={"start_year": 2024, "end_year": 2023})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_unreachable_database_returns_503(tmp_path: Path, monkeypatch):
    # Point DATABASE_URL at a postgres host that will never accept connections.
    monkeypatch.setenv("DATABASE_URL", "postgresql://user:pass@127.0.0.1:1/nonexistent")
    settings = build_settings(project_root=tmp_path)

    app = create_app(settings=settings)
    with TestClient(app, raise_server_exceptions=False) as client:
        response = client.get("/companies")

    assert response.status_code == 503
    assert response.json()["error"]["code"] == "service_unavailable"


def test_service_failure_returns_503(client: TestClient):
    app = client.app
    original_service = app.state.read_service

    class BrokenService:
        engine = original_service.engine

        def list_companies(self, **_: object):
            raise RuntimeError("database is down")

    app.state.read_service = BrokenService()
    try:
        with TestClient(app, raise_server_exceptions=False) as broken_client:
            response = broken_client.get("/companies")
    finally:
        app.state.read_service = original_service

    assert response.status_code == 503
    assert response.json()["error"]["code"] == "service_unavailable"


def test_companies_reject_invalid_page(client: TestClient):
    response = client.get("/companies", params={"page": 0})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "validation_error"


def test_companies_reject_invalid_page_size(client: TestClient):
    response = client.get("/companies", params={"page_size": 101})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "validation_error"


# ---------------------------------------------------------------------------
# Summary endpoint
# ---------------------------------------------------------------------------


def test_company_summary_returns_blocks(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2023,2024"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["cd_cvm"] == 9512
    assert payload["years"] == [2023, 2024]
    assert len(payload["blocks"]) >= 1
    assert payload["blocks"][0]["stmt_type"] in {"DRE", "BPA", "BPP", "DFC"}


def test_company_summary_block_columns_contain_required_fields(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2023,2024"})

    assert response.status_code == 200
    for block in response.json()["blocks"]:
        cols = block["table"]["columns"]
        assert "CD_CONTA" in cols
        assert "LABEL" in cols
        assert "IS_SUBTOTAL" in cols


def test_company_summary_is_subtotal_propagation(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2023"})

    assert response.status_code == 200
    dre_block = next((b for b in response.json()["blocks"] if b["stmt_type"] == "DRE"), None)
    assert dre_block is not None
    row_3_01 = next((r for r in dre_block["table"]["rows"] if r["CD_CONTA"] == "3.01"), None)
    assert row_3_01 is not None
    assert row_3_01["IS_SUBTOTAL"] is True


def test_company_summary_empty_blocks_when_no_data(client: TestClient):
    # year 1990 has no data for PETROBRAS → blocks should be []
    response = client.get("/companies/9512/summary", params={"years": "1990"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["blocks"] == []


def test_company_summary_404_unknown_company(client: TestClient):
    response = client.get("/companies/999999/summary", params={"years": "2024"})

    assert response.status_code == 404
    assert response.json()["error"]["code"] == "not_found"


def test_company_summary_422_invalid_years(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2024,foo"})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_summary_422_duplicate_years(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2024,2024"})

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "invalid_request"


def test_company_summary_years_sorted_stable(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2024,2023"})

    assert response.status_code == 200
    assert response.json()["years"] == [2023, 2024]


def test_company_summary_single_year_columns(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2024"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["years"] == [2024]
    assert len(payload["blocks"]) >= 1
    first_block = payload["blocks"][0]
    assert "2024" in first_block["table"]["columns"]
    assert "2023" not in first_block["table"]["columns"]


def test_company_summary_blocks_have_non_empty_titles(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2023"})

    assert response.status_code == 200
    for block in response.json()["blocks"]:
        assert isinstance(block["title"], str) and len(block["title"]) > 0


def test_non_targeted_endpoints_do_not_expose_api_cache_headers(client: TestClient):
    response = client.get("/companies/9512/summary", params={"years": "2023,2024"})

    assert response.status_code == 200
    assert "cache-control" not in response.headers


# ── Regressao: /years exclui anos com apenas dados trimestrais ITR ─────────────
# A seed do banco inclui linhas ITR para PETROBRAS (REPORT_YEAR=2025,
# PERIOD_LABEL="1Q25"). Sem o filtro PERIOD_LABEL = CAST(REPORT_YEAR AS TEXT)
# o endpoint retornaria [2023, 2024, 2025] e o Comparar usaria referenceYear=2025,
# cujas colunas nao existem no KPI bundle (KPI engine so usa dados anuais).
# Isso causaria todos os valores "-" na tabela de comparacao.

def test_company_years_excludes_itr_only_years(client: TestClient):
    """Garante que /years retorna apenas anos com DFP (PERIOD_LABEL == ano)."""
    response = client.get("/companies/9512/years")

    assert response.status_code == 200
    years = response.json()
    # 2025 so tem ITR (PERIOD_LABEL="1Q25") no seed — nao deve aparecer
    assert 2025 not in years
    # anos com DFP completo devem permanecer
    assert 2023 in years
    assert 2024 in years


# ── CORS: ALLOWED_ORIGINS configuravel via env var ─────────────────────────────


def test_cors_default_allows_localhost(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    """Sem ALLOWED_ORIGINS, apenas http://localhost:3000 e aceito."""
    monkeypatch.delenv("ALLOWED_ORIGINS", raising=False)
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'cors_test.db'}")
    settings = build_settings()
    test_app = create_app(settings=settings)
    with TestClient(test_app) as c:
        response = c.options(
            "/health",
            headers={"Origin": "http://localhost:3000", "Access-Control-Request-Method": "GET"},
        )
    assert response.status_code == 200
    assert response.headers.get("access-control-allow-origin") == "http://localhost:3000"


def test_cors_env_var_allows_multiple_origins(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    """ALLOWED_ORIGINS com varios valores separados por virgula — todos aceitos."""
    monkeypatch.setenv("ALLOWED_ORIGINS", "https://app.vercel.app,https://staging.vercel.app")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'cors_multi.db'}")
    settings = build_settings()
    test_app = create_app(settings=settings)
    with TestClient(test_app) as c:
        for origin in ("https://app.vercel.app", "https://staging.vercel.app"):
            resp = c.options(
                "/health",
                headers={"Origin": origin, "Access-Control-Request-Method": "GET"},
            )
            assert resp.headers.get("access-control-allow-origin") == origin, (
                f"Origem {origin} nao foi aceita pelo CORS"
            )


# ── Sentry: inicializacao condicional via SENTRY_DSN ──────────────────────────


def test_sentry_skipped_when_dsn_is_absent(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    """Sem SENTRY_DSN, o Sentry nao deve ser inicializado."""
    import sentry_sdk

    monkeypatch.delenv("SENTRY_DSN", raising=False)
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'sentry_off.db'}")
    settings = build_settings()
    create_app(settings=settings)
    assert sentry_sdk.get_client().dsn is None


def test_sentry_initialized_when_dsn_is_present(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    """Com SENTRY_DSN valido, o Sentry deve ser inicializado com o environment correto."""
    import sentry_sdk

    fake_dsn = "https://pub@o0.ingest.sentry.io/0"
    monkeypatch.setenv("SENTRY_DSN", fake_dsn)
    monkeypatch.setenv("SENTRY_ENVIRONMENT", "staging")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'sentry_on.db'}")
    settings = build_settings()
    create_app(settings=settings)
    client = sentry_sdk.get_client()
    assert client.dsn == fake_dsn
    assert client.options["environment"] == "staging"
