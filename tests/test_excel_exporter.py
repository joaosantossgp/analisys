import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.excel_exporter import ExcelExporter, build_excel_filename


def _statement(rows: list[dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _minimal_kpis() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "CATEGORIA": "Rentabilidade",
            "KPI_ID": "MG_EBIT",
            "KPI_NOME": "Margem EBIT",
            "FORMULA": "EBIT / Receita",
            "IS_PLACEHOLDER": False,
            "FORMAT_TYPE": "pct",
            "HIGHER_IS_BETTER": True,
            "UNIDADE": "%",
            "2024": 0.2,
            "DELTA_YOY": None,
            "DELTA_YOY_PCT": None,
        }
    ])


def _sample_company() -> dict[str, object]:
    return {
        "company_name": "Empresa Exemplo",
        "cd_cvm": 1234,
        "ticker_b3": "EXMP3",
        "cnpj": "00.000.000/0001-00",
        "setor_cvm": "Teste",
        "setor_analitico": "Teste",
    }


def _sample_statements() -> dict[str, pd.DataFrame]:
    return {
        "DRE": _statement([
            {"CD_CONTA": "3.01", "DS_CONTA": "Receita", "STANDARD_NAME": "", "LINE_ID_BASE": "3.01", "2024": 100.0, "1Q25": 30.0},
            {"CD_CONTA": "3.03", "DS_CONTA": "Resultado Bruto", "STANDARD_NAME": "", "LINE_ID_BASE": "3.03", "2024": 50.0, "1Q25": 12.0},
            {"CD_CONTA": "3.05", "DS_CONTA": "EBIT", "STANDARD_NAME": "", "LINE_ID_BASE": "3.05", "2024": 20.0, "1Q25": 5.0},
            {"CD_CONTA": "3.07", "DS_CONTA": "LAIR", "STANDARD_NAME": "", "LINE_ID_BASE": "3.07", "2024": 16.0, "1Q25": 4.0},
            {"CD_CONTA": "3.11", "DS_CONTA": "Lucro Liquido", "STANDARD_NAME": "", "LINE_ID_BASE": "3.11", "2024": 12.0, "1Q25": 3.0},
        ]),
        "BPA": _statement([
            {"CD_CONTA": "1", "DS_CONTA": "Ativo Total", "STANDARD_NAME": "", "LINE_ID_BASE": "1", "2024": 210.0, "1Q25": 220.0},
            {"CD_CONTA": "1.01", "DS_CONTA": "Ativo Circulante", "STANDARD_NAME": "", "LINE_ID_BASE": "1.01", "2024": 80.0, "1Q25": 82.0},
            {"CD_CONTA": "1.01.01", "DS_CONTA": "Caixa", "STANDARD_NAME": "", "LINE_ID_BASE": "1.01.01", "2024": 35.0, "1Q25": 34.0},
            {"CD_CONTA": "1.01.02", "DS_CONTA": "Aplicacoes", "STANDARD_NAME": "", "LINE_ID_BASE": "1.01.02", "2024": 10.0, "1Q25": 12.0},
            {"CD_CONTA": "1.02", "DS_CONTA": "Ativo Nao Circulante", "STANDARD_NAME": "", "LINE_ID_BASE": "1.02", "2024": 130.0, "1Q25": 138.0},
            {"CD_CONTA": "1.02.03", "DS_CONTA": "Imobilizado", "STANDARD_NAME": "", "LINE_ID_BASE": "1.02.03", "2024": 90.0, "1Q25": 95.0},
        ]),
        "BPP": _statement([
            {"CD_CONTA": "2", "DS_CONTA": "Passivo Total", "STANDARD_NAME": "", "LINE_ID_BASE": "2", "2024": 210.0, "1Q25": 220.0},
            {"CD_CONTA": "2.01", "DS_CONTA": "Passivo Circulante", "STANDARD_NAME": "", "LINE_ID_BASE": "2.01", "2024": 90.0, "1Q25": 95.0},
            {"CD_CONTA": "2.01.01", "DS_CONTA": "Obrigacoes", "STANDARD_NAME": "", "LINE_ID_BASE": "2.01.01", "2024": 12.0, "1Q25": 13.0},
            {"CD_CONTA": "2.01.04", "DS_CONTA": "Emprestimos CP", "STANDARD_NAME": "", "LINE_ID_BASE": "2.01.04", "2024": 25.0, "1Q25": 27.0},
            {"CD_CONTA": "2.02", "DS_CONTA": "Passivo Nao Circulante", "STANDARD_NAME": "", "LINE_ID_BASE": "2.02", "2024": 60.0, "1Q25": 61.0},
            {"CD_CONTA": "2.02.01", "DS_CONTA": "Emprestimos LP", "STANDARD_NAME": "", "LINE_ID_BASE": "2.02.01", "2024": 40.0, "1Q25": 42.0},
            {"CD_CONTA": "2.03", "DS_CONTA": "Patrimonio Liquido", "STANDARD_NAME": "", "LINE_ID_BASE": "2.03", "2024": 60.0, "1Q25": 64.0},
            {"CD_CONTA": "2.03.01", "DS_CONTA": "Capital Social", "STANDARD_NAME": "", "LINE_ID_BASE": "2.03.01", "2024": 30.0, "1Q25": 30.0},
            {"CD_CONTA": "2.03.04", "DS_CONTA": "Reservas de Lucros", "STANDARD_NAME": "", "LINE_ID_BASE": "2.03.04", "2024": 20.0, "1Q25": 24.0},
        ]),
        "DFC": _statement([
            {"CD_CONTA": "6.01", "DS_CONTA": "Fluxo Operacional", "STANDARD_NAME": "", "LINE_ID_BASE": "6.01", "2024": 25.0, "1Q25": 7.0},
            {"CD_CONTA": "6.01.01", "DS_CONTA": "Caixa Gerado nas Operacoes", "STANDARD_NAME": "", "LINE_ID_BASE": "6.01.01", "2024": 28.0, "1Q25": 8.0},
            {"CD_CONTA": "6.01.01.03", "DS_CONTA": "Depreciacao e Amortizacao", "STANDARD_NAME": "", "LINE_ID_BASE": "6.01.01.03", "2024": 2.0, "1Q25": 0.5},
            {"CD_CONTA": "6.02", "DS_CONTA": "Fluxo Investimento", "STANDARD_NAME": "", "LINE_ID_BASE": "6.02", "2024": -12.0, "1Q25": -3.0},
            {"CD_CONTA": "6.02.01", "DS_CONTA": "Aquisicao de Imobilizado", "STANDARD_NAME": "", "LINE_ID_BASE": "6.02.01", "2024": -10.0, "1Q25": -2.0},
            {"CD_CONTA": "6.03", "DS_CONTA": "Fluxo Financiamento", "STANDARD_NAME": "", "LINE_ID_BASE": "6.03", "2024": -6.0, "1Q25": -2.0},
            {"CD_CONTA": "6.03.01", "DS_CONTA": "Captacao de Emprestimos", "STANDARD_NAME": "", "LINE_ID_BASE": "6.03.01", "2024": 9.0, "1Q25": 2.0},
            {"CD_CONTA": "6.05", "DS_CONTA": "Aumento (Reducao) de Caixa e Equivalentes", "STANDARD_NAME": "", "LINE_ID_BASE": "6.05", "2024": 7.0, "1Q25": 2.0},
            {"CD_CONTA": "6.05.02", "DS_CONTA": "Saldo Final de Caixa e Equivalentes", "STANDARD_NAME": "", "LINE_ID_BASE": "6.05.02", "2024": 35.0, "1Q25": 34.0},
        ]),
        "DVA": _statement([
            {"CD_CONTA": "7.08", "DS_CONTA": "Valor Adicionado", "STANDARD_NAME": "", "LINE_ID_BASE": "7.08", "2024": 10.0},
        ]),
    }


def test_excel_exporter_includes_geral_sheet_in_expected_order():
    exporter = ExcelExporter(
        company_info=_sample_company(),
        statements=_sample_statements(),
        kpis_df=_minimal_kpis(),
        extra_sheets=["DVA"],
    )

    workbook_bytes = exporter.export()
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))

    assert wb.sheetnames == ["CAPA", "GERAL", "KPIs", "DRE", "BPA", "BPP", "DFC", "DVA", "METADADOS"]


def test_excel_exporter_geral_sheet_preserves_periods_and_expanded_rows():
    exporter = ExcelExporter(
        company_info=_sample_company(),
        statements=_sample_statements(),
        kpis_df=_minimal_kpis(),
        extra_sheets=[],
    )

    workbook_bytes = exporter.export()
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    geral = wb["GERAL"]
    dre = wb["DRE"]

    first_col = [geral.cell(row, 1).value for row in range(1, geral.max_row + 1) if geral.cell(row, 1).value is not None]

    assert first_col[0] == "DRE — Resumo Condensado"
    assert "BPA — Resumo Condensado" in first_col
    assert "BPP — Resumo Condensado" in first_col
    assert "DFC — Resumo Condensado" in first_col
    assert first_col.index("BPA — Resumo Condensado") < first_col.index("BPP — Resumo Condensado") < first_col.index("DFC — Resumo Condensado")

    assert "1.01.02" in first_col
    assert "2.01.01" in first_col
    assert "6.01.01" in first_col
    assert "6.01.01.03" in first_col
    assert "6.02.01" in first_col
    assert "6.03.01" in first_col
    assert "6.05" in first_col
    assert "6.05.02" in first_col

    assert geral["A2"].value == "CD_CONTA"
    assert geral["B2"].value == "LINHA"
    assert geral["C2"].value == "2024"
    assert geral["D2"].value == "1Q25"

    assert dre["A1"].value == "CD_CONTA"
    assert dre["B2"].value == "Receita"


def test_build_excel_filename_uses_ticker_and_export_date():
    filename = build_excel_filename(
        _sample_company(),
        generated_at=datetime(2026, 4, 9, 14, 30),
    )

    assert filename == "EXMP3_20260409.xlsx"


def test_build_excel_filename_falls_back_to_cvm_code_without_ticker():
    filename = build_excel_filename(
        {
            **_sample_company(),
            "ticker_b3": None,
        },
        generated_at=datetime(2026, 4, 9, 14, 30),
    )

    assert filename == "cvm1234_20260409.xlsx"
